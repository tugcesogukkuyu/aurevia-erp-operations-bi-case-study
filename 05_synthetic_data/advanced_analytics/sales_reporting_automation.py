"""
===============================================================================
Aurevia ERP Operations & BI Dashboard
Sales Reporting Automation

Purpose
-------
This script creates a refreshable Excel-based management reporting package from
the SQL reporting views used in the advanced dashboard extension.

It supports the Power BI page:
07 - Sales Operations Command Center

Business Purpose
----------------
Automate the extraction of sales, profitability, receivables, and operational
alert outputs so that management can review the same reporting layer used by
Power BI.

Technology Stack
----------------
- pandas: SQL output processing
- SQLAlchemy / pyodbc: SQL Server connection
- openpyxl: Excel report generation
- logging: execution traceability

Input SQL Views
---------------
- vw_Page07_MonthlySalesCommandTrend
- vw_Page07_TopCustomersByRevenue
- vw_Page07_ProductCategoryRevenueRanking
- vw_Page07_ReceivablesOpenBalanceRisk
- vw_Page07_OperationalAlerts
- vw_Page07_TopOverdueCustomerExposure

Output
------
outputs/aurevia_sales_operations_report.xlsx

===============================================================================
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine


# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# -----------------------------------------------------------------------------
# Logging
# -----------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)

logger = logging.getLogger("aurevia_sales_reporting_automation")


# -----------------------------------------------------------------------------
# SQL Connection
# -----------------------------------------------------------------------------

def get_sql_engine():
    """
    Creates a SQL Server connection engine.

    Environment variables:
        AUREVIA_SQL_SERVER
        AUREVIA_SQL_DATABASE
        AUREVIA_SQL_USERNAME
        AUREVIA_SQL_PASSWORD
    """

    server = os.getenv("AUREVIA_SQL_SERVER")
    database = os.getenv("AUREVIA_SQL_DATABASE")
    username = os.getenv("AUREVIA_SQL_USERNAME")
    password = os.getenv("AUREVIA_SQL_PASSWORD")

    if not all([server, database, username, password]):
        raise RuntimeError(
            "SQL connection variables are missing. "
            "Set AUREVIA_SQL_SERVER, AUREVIA_SQL_DATABASE, "
            "AUREVIA_SQL_USERNAME, and AUREVIA_SQL_PASSWORD."
        )

    connection_string = (
        "mssql+pyodbc://"
        f"{username}:{password}@{server}/{database}"
        "?driver=ODBC+Driver+18+for+SQL+Server"
        "&TrustServerCertificate=yes"
    )

    return create_engine(connection_string, fast_executemany=True)


# -----------------------------------------------------------------------------
# Query Definitions
# -----------------------------------------------------------------------------

REPORT_QUERIES: Dict[str, str] = {
    "KPI_Reconciliation": """
        SELECT
            'Total Revenue' AS Metric,
            ROUND(SUM(TotalRevenue), 2) AS Value
        FROM vw_Page07_MonthlySalesCommandTrend

        UNION ALL

        SELECT
            'Gross Profit' AS Metric,
            ROUND(SUM(GrossProfit), 2) AS Value
        FROM vw_Page07_MonthlySalesCommandTrend

        UNION ALL

        SELECT
            'Open Balance' AS Metric,
            ROUND(SUM(OpenBalance), 2) AS Value
        FROM vw_Page07_ReceivablesOpenBalanceRisk;
    """,

    "Monthly_Trend": """
        SELECT
            YearMonth,
            TotalRevenue,
            GrossProfit,
            GrossMarginPercent
        FROM vw_Page07_MonthlySalesCommandTrend
        ORDER BY YearMonth;
    """,

    "Top_Customers": """
        SELECT
            RevenueRank,
            CustomerName,
            CustomerSegment,
            Region,
            SalesChannel,
            TotalRevenue,
            GrossProfit,
            OpenBalance,
            RevenueSharePercent
        FROM vw_Page07_TopCustomersByRevenue
        ORDER BY RevenueRank;
    """,

    "Product_Category_Revenue": """
        SELECT
            ProductCategory,
            TotalRevenue,
            GrossProfit,
            GrossMarginPercent,
            TotalQuantitySold,
            ProductCount
        FROM vw_Page07_ProductCategoryRevenueRanking
        ORDER BY TotalRevenue DESC;
    """,

    "Receivables_Aging": """
        SELECT
            AgingBucket,
            AgingSortOrder,
            OpenBalance,
            InvoiceCount,
            OpenBalanceSharePercent
        FROM vw_Page07_ReceivablesOpenBalanceRisk
        ORDER BY AgingSortOrder;
    """,

    "Operational_Alerts": """
        SELECT
            AlertType,
            AlertValue,
            AlertSeverity,
            OwnerArea
        FROM vw_Page07_OperationalAlerts
        ORDER BY
            CASE AlertSeverity
                WHEN 'High' THEN 1
                WHEN 'Medium' THEN 2
                WHEN 'Low' THEN 3
                ELSE 99
            END,
            AlertType;
    """,

    "Top_Overdue_Customers": """
        SELECT
            CustomerName,
            CustomerSegment,
            OpenBalance,
            MaxDaysOverdue
        FROM vw_Page07_TopOverdueCustomerExposure
        ORDER BY OpenBalance DESC;
    """,
}


# -----------------------------------------------------------------------------
# Data Extraction
# -----------------------------------------------------------------------------

def extract_report_tables() -> Dict[str, pd.DataFrame]:
    """
    Extracts all report tables from SQL Server.
    """

    logger.info("Starting SQL extraction for sales operations report.")

    engine = get_sql_engine()

    tables: Dict[str, pd.DataFrame] = {}

    for sheet_name, query in REPORT_QUERIES.items():
        logger.info("Extracting: %s", sheet_name)
        tables[sheet_name] = pd.read_sql(query, engine)
        logger.info("Rows extracted for %s: %s", sheet_name, len(tables[sheet_name]))

    logger.info("SQL extraction completed.")

    return tables


# -----------------------------------------------------------------------------
# Excel Formatting
# -----------------------------------------------------------------------------

def auto_size_columns(worksheet) -> None:
    """
    Auto-sizes worksheet columns based on cell content length.
    """

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass

        adjusted_width = min(max_length + 2, 40)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def format_header_row(worksheet) -> None:
    """
    Applies consistent header styling.
    """

    header_fill = PatternFill(
        start_color="17324D",
        end_color="17324D",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_basic_formatting(writer) -> None:
    """
    Applies basic formatting to all worksheets.
    """

    workbook = writer.book

    for worksheet in workbook.worksheets:
        format_header_row(worksheet)
        auto_size_columns(worksheet)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")

        worksheet.freeze_panes = "A2"


# -----------------------------------------------------------------------------
# Excel Report Writer
# -----------------------------------------------------------------------------

def write_excel_report(tables: Dict[str, pd.DataFrame]) -> Path:
    """
    Writes extracted SQL tables into a structured Excel workbook.
    """

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_file = OUTPUT_DIR / f"aurevia_sales_operations_report_{timestamp}.xlsx"

    logger.info("Writing Excel report: %s", output_file)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            safe_sheet_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

        apply_basic_formatting(writer)

    logger.info("Excel report created successfully.")

    return output_file


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main() -> None:
    """
    Executes the sales reporting automation pipeline.
    """

    tables = extract_report_tables()
    output_file = write_excel_report(tables)

    logger.info("Sales reporting automation completed.")
    logger.info("Output file: %s", output_file)


if __name__ == "__main__":
    main()